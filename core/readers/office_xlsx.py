"""Spreadsheets -> core.ir.Document, via openpyxl.

**One worksheet is one page, while it fits in one.** A sheet boundary is real --
the file declares it, sheets are named and ordered, and flattening them into a
character count would throw away the only structure a spreadsheet has. So a
workbook of ordinary sheets reports `pagination: native` and one page each.

A sheet too large for one page is divided, and then the workbook reports
`pagination: synthetic`, because those boundaries are this server's rather than
the file's. `sheet_pages` maps each sheet name to the pages it occupies, so the
caller can still ask for a sheet by name.

That case is the common one on real data, not an edge: Car_Insurance_Fraud.xlsx
in the corpus is a SINGLE sheet worth 312,000 tokens. Held as one page it was
unreachable -- extract refused it for exceeding the budget, and the refusal's
hint named a page range that could not be narrowed, because there was only
page 1. A refusal a caller cannot act on is worse than no tool at all.

`read_only=True` and `data_only=True`, both load-bearing:

    read_only   openpyxl's normal mode builds a cell object per cell. A
                200,000-row export is then a couple of gigabytes of Python
                objects against a 1 GiB container. Read-only streams rows.
    data_only   returns the cached VALUE of a formula rather than "=SUM(A1:A9)".
                A caller asking a document server what a spreadsheet says wants
                the number. Where Excel never cached one -- a file written by a
                library rather than by Excel -- the value is None, and this
                reports the empty cell rather than inventing a zero.

A sheet's used range is what openpyxl reports as its dimensions, which are
frequently far larger than its data: a stray format applied to a whole column
makes `max_row` a million. Trailing empty rows and columns are trimmed, and the
count reported is the real one.
"""

from __future__ import annotations

from core.ir import Block, Document, Page, Span
from core.readers import _flow
from core.selection import format_pages

BODY_SIZE = 12.0

# Rows read from one sheet before stopping. A spreadsheet is not a document and
# a caller who needs all 200,000 rows of one wants the data servers, not this.
# The response says how many were read and how many exist.
MAX_ROWS_PER_SHEET = 5000


def open_document(path: str, password: str = "") -> Document:
    """Read a workbook. One page per worksheet, values not formulas."""
    import openpyxl

    src = _flow.guard_size(path)
    try:
        book = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    except Exception as exc:
        from core.readers import ReaderError

        message = str(exc)
        if "encrypted" in message.lower() or "not a zip" in message.lower():
            raise ReaderError(
                f"{src.name} could not be opened; it may be password-protected or not a .xlsx.",
                "Excel's own encryption is not readable here. Save an unprotected copy first.",
            ) from exc
        raise ReaderError(
            f"{src.name} could not be read as a spreadsheet: {message}",
            "Check the file opens in Excel. A .xls (not .xlsx) needs convert(to='pdf') first.",
        ) from exc

    pages: list[Page] = []
    truncated: list[str] = []
    titles: list[str] = []
    sheet_pages: dict[str, str] = {}
    split = False
    try:
        for sheet in book.worksheets:
            titles.append(sheet.title)
            rows, total = _rows(sheet)
            if total > len(rows):
                truncated.append(sheet.title)

            # A sheet is one page only while it FITS in one. A real export --
            # Car_Insurance_Fraud.xlsx in the corpus -- is a single sheet worth
            # 312,000 tokens, and "one sheet, one page" made that a page no
            # caller could ever read: extract refuses it for being over the
            # budget, and the hint it gives names a page range that cannot be
            # narrowed, because there is only page 1. A refusal the caller
            # cannot act on is worse than no tool.
            first = len(pages) + 1
            made = _flow.paginate(_blocks(sheet.title, rows, total), "native")
            split = split or len(made) > 1
            pages.extend(made)
            sheet_pages[sheet.title] = format_pages(list(range(first, len(pages) + 1)))
    finally:
        # Collected inside the loop, not read back afterwards: a read_only
        # workbook's worksheet objects are backed by the archive this closes,
        # so `book.worksheets` after close() is a use-after-close.
        book.close()

    meta = {
        "sheets": titles,
        "sheet_count": len(titles),
        # Which pages hold which sheet. Without this, a workbook whose second
        # sheet starts on page 7 gives the caller no way to ask for it -- the
        # sheet names are in the response and the page numbers are in the
        # response and nothing connects them.
        "sheet_pages": sheet_pages,
        # Honest when a sheet had to be divided: those page boundaries are this
        # server's, not the workbook's, and the caller is told which it is.
        "pagination": "synthetic" if split else "native",
    }
    if truncated:
        meta["truncated_sheets"] = truncated
        meta["max_rows_per_sheet"] = MAX_ROWS_PER_SHEET
    return _flow.paged(str(src), "xlsx", pages, meta)


def _rows(sheet) -> tuple[list[list[str]], int]:
    """Non-empty rows, trimmed of trailing empty columns, and the real total.

    The total counts rows actually iterated, not `sheet.max_row`: a sheet whose
    formatting extends to row 1,048,576 reports that as its maximum, and using
    it would make every response claim a million rows for a file with nine.
    """
    out: list[list[str]] = []
    total = 0
    for raw in sheet.iter_rows(values_only=True):
        cells = ["" if value is None else str(value) for value in raw]
        while cells and not cells[-1].strip():
            cells.pop()
        if not cells:
            continue
        total += 1
        if len(out) < MAX_ROWS_PER_SHEET:
            out.append(cells)
    return out, total


def _blocks(title: str, rows: list[list[str]], total: int) -> list[Block]:
    """A heading for the sheet name, then the sheet as one table block.

    The sheet name is a heading so `outline` lists the sheets, which is the
    navigation a workbook actually has.
    """
    blocks: list[Block] = [_flow.block(title, "heading", "native", 22.0)]
    if not rows:
        return blocks
    flat = "\n".join("\t".join(cell for cell in row) for row in rows)
    blocks.append(Block(kind="table", spans=[Span(text=flat, size=BODY_SIZE)], bbox=None, basis="native", rows=rows))
    if total > len(rows):
        blocks.append(
            _flow.block(
                f"[{total - len(rows)} further row(s) not read; this reader stops at {MAX_ROWS_PER_SHEET}]",
                "caption",
                "native",
                BODY_SIZE,
            )
        )
    return blocks


def page_tables(doc: Document, numbers: list[int]) -> list[dict]:
    from core.readers.html import page_tables as _shared

    return _shared(doc, numbers)


def bookmarks(doc: Document) -> list[dict]:
    """One entry per sheet. A workbook's outline is its sheet names."""
    out: list[dict] = []
    for number in sorted(doc.pages):
        for block in doc.pages[number].blocks:
            if block.kind == "heading":
                out.append({"level": 1, "title": block.text.strip(), "page": number, "basis": "native"})
                break
    return out


def probe_extras(doc: Document) -> dict:
    keys = ("sheets", "sheet_count", "sheet_pages", "truncated_sheets", "max_rows_per_sheet")
    return {key: doc.meta[key] for key in keys if key in doc.meta}


close_document = _flow.close_document
load_page = _flow.load_page
load_page_words = _flow.load_page
